import json
import os
from typing import Any

import unittest
from unittest.mock import mock_open, patch
import openpyxl
from typing import List, Dict
from src.utils import read_transactions_from_json
from src.utils import read_transactions_from_csv, read_transactions_from_xlsx


# Вспомогательная функция для создания временного JSON-файла
def create_temp_json_file(data: Any, filename: str) -> str:
    file_path = os.path.join(os.path.dirname(__file__), filename)
    with open(file_path, 'w', encoding='utf-8') as file:
        json.dump(data, file)
    return file_path


def test_read_transactions_from_json_success():
    data: List[Dict[str, Any]] = [{"id": 1, "amount": 100}, {"id": 2, "amount": 200}]
    file_path = create_temp_json_file(data, 'test_data.json')
    result = read_transactions_from_json(file_path)
    assert result == data
    os.remove(file_path)


def test_read_transactions_from_json_file_not_found():
    result = read_transactions_from_json('non_existing_file.json')
    assert result == []


def test_read_transactions_from_json_invalid_format():
    # Неверный формат: не список
    data: Dict[str, int] = {"id": 1, "amount": 100}
    file_path = create_temp_json_file(data, 'test_data_invalid.json')
    result = read_transactions_from_json(file_path)
    assert result == []
    os.remove(file_path)


def test_read_transactions_from_json_empty_list():
    # Верный формат: пустой список
    data: List[Dict[str, Any]] = []
    file_path = create_temp_json_file(data, 'test_data_empty.json')
    result = read_transactions_from_json(file_path)
    assert result == data
    os.remove(file_path)


# Импортируем функции, которые необходимо протестировать


class TestReadTransactions(unittest.TestCase):

    @patch("builtins.open", new_callable=mock_open, read_data="id,name,amount\n1,Alice,100\n2,Bob,150")
    @patch("os.path.dirname", return_value="")  # Подмена текущего каталога пустой строкой
    def test_read_transactions_from_csv(self, mock_dirname, mock_file):
        expected_result: List[Dict[str, str]] = [
            {"id": "1", "name": "Alice", "amount": "100"},
            {"id": "2", "name": "Bob", "amount": "150"},
        ]

        result = read_transactions_from_csv("transactions.csv")
        self.assertEqual(result, expected_result)

    @patch("openpyxl.load_workbook")
    @patch("os.path.dirname", return_value="")  # Подмена текущего каталога пустой строкой
    def test_read_transactions_from_xlsx(self, mock_dirname, mock_load_workbook):
        # Создание фиктивной рабочей книги и листа
        mock_workbook = openpyxl.Workbook()
        mock_sheet = mock_workbook.active
        mock_sheet.append(["id", "name", "amount"])
        mock_sheet.append([1, "Alice", 100])
        mock_sheet.append([2, "Bob", 150])

        mock_load_workbook.return_value = mock_workbook

        expected_result: List[Dict[str, str]] = [
            {"id": 1, "name": "Alice", "amount": 100},
            {"id": 2, "name": "Bob", "amount": 150},
        ]

        result = read_transactions_from_xlsx("transactions.xlsx")
        self.assertEqual(result, expected_result)
